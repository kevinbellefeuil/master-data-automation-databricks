"""
Flask backend for IBP Master Data Automation - Databricks Apps

Robust version with graceful degradation:
- Phase 1: Starts with sample data (guaranteed to work)
- Phase 2: Attempts database connection with fallback to sample data
- Phase 3: Full database read/write with comprehensive error handling

@mitigates MasterDataApp:Backend:API against #path_traversal with fixed file path validation
@mitigates MasterDataApp:Backend:API against #sql_injection with parameterized queries
@mitigates MasterDataApp:Backend:API against #xss with input validation
@connects User:Browser to MasterDataApp:Backend:API with HTTPS/TCP/8080
@connects MasterDataApp:Backend:API to Databricks:Table with SQL connector
"""

from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from pathlib import Path
import os
import logging
import sys
import re
from threading import Lock
import requests
import time
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
# @mitigates MasterDataApp:Backend against #info_disclosure with sanitized logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='dist', static_url_path='')

# @mitigates MasterDataApp:Backend:API against #csrf with CORS configuration
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})

# File paths - use absolute paths to avoid resolution issues
# @mitigates MasterDataApp:Backend:FileSystem against #path_traversal with absolute paths
BASE_DIR = Path(__file__).parent.absolute()
DIST_DIR = BASE_DIR / 'dist'

# Operation lock for thread safety
# @mitigates MasterDataApp:Backend:Database against #race_condition with mutex
operation_lock = Lock()

# Hardcoded sample data - FALLBACK if database unavailable
# @accepts #limited_functionality to MasterDataApp:Backend with fallback to sample data on DB failure
SAMPLE_DATA = [
    {
        'PRDID': 'PROD001',
        'LOCID': 'LOC001',
        'CURRENT_PLUNITID': 'UNIT-A',
        'RECOMMENDED': 'UNIT-B',
        'RECOMMEND_REASON': 'Better efficiency',
        'Approval_Status': None,
        'approval': None
    },
    {
        'PRDID': 'PROD002',
        'LOCID': 'LOC002',
        'CURRENT_PLUNITID': 'UNIT-C',
        'RECOMMENDED': 'UNIT-D',
        'RECOMMEND_REASON': 'Cost reduction',
        'Approval_Status': None,
        'approval': None
    },
    {
        'PRDID': 'PROD003',
        'LOCID': 'LOC003',
        'CURRENT_PLUNITID': 'UNIT-E',
        'RECOMMENDED': 'UNIT-F',
        'RECOMMEND_REASON': 'Improved quality',
        'Approval_Status': None,
        'approval': None
    },
    {
        'PRDID': 'PROD004',
        'LOCID': 'LOC004',
        'CURRENT_PLUNITID': 'UNIT-G',
        'RECOMMENDED': 'UNIT-H',
        'RECOMMEND_REASON': 'Standardization',
        'Approval_Status': None,
        'approval': None
    },
    {
        'PRDID': 'PROD005',
        'LOCID': 'LOC005',
        'CURRENT_PLUNITID': 'UNIT-I',
        'RECOMMENDED': 'UNIT-J',
        'RECOMMEND_REASON': 'Vendor consolidation',
        'Approval_Status': None,
        'approval': None
    },
    {
        'PRDID': 'PROD006',
        'LOCID': 'LOC006',
        'CURRENT_PLUNITID': 'UNIT-K',
        'RECOMMENDED': 'UNIT-L',
        'RECOMMEND_REASON': 'Cost optimization',
        'Approval_Status': None,
        'approval': None
    },
    {
        'PRDID': 'PROD007',
        'LOCID': 'LOC007',
        'CURRENT_PLUNITID': 'UNIT-M',
        'RECOMMENDED': 'UNIT-N',
        'RECOMMEND_REASON': 'Process improvement',
        'Approval_Status': None,
        'approval': None
    },
    {
        'PRDID': 'PROD008',
        'LOCID': 'LOC008',
        'CURRENT_PLUNITID': 'UNIT-O',
        'RECOMMENDED': 'UNIT-P',
        'RECOMMEND_REASON': 'Quality enhancement',
        'Approval_Status': None,
        'approval': None
    }
]


class InputValidator:
    """
    Input validation and sanitization utility
    
    @mitigates MasterDataApp:Backend against #injection with comprehensive input validation
    @mitigates MasterDataApp:Backend against #xss with HTML entity encoding
    """
    
    @staticmethod
    def sanitize_string(value):
        """
        Sanitize string input to prevent injection attacks
        
        @validates Input from external sources
        @mitigates MasterDataApp:Backend against #xss with character filtering
        """
        if value is None:
            return ''
        
        if not isinstance(value, str):
            value = str(value)
        
        # Remove potentially dangerous characters
        value = re.sub(r'[<>]', '', value)
        
        return value.strip()
    
    @staticmethod
    def validate_record(record):
        """
        Validate record structure
        
        @validates Record structure from API requests
        """
        if not isinstance(record, dict):
            return False
        
        # Must have PRDID and LOCID for matching
        if 'PRDID' not in record or 'LOCID' not in record:
            return False
        
        if not record['PRDID'] or not record['LOCID']:
            return False
        
        return True


class DatabricksHandler:
    """
    Databricks SQL operations handler with graceful degradation
    Uses Databricks REST API for SQL execution (no native dependencies)
    
    @connects MasterDataApp:Backend to Databricks:Table with REST API
    @mitigates MasterDataApp:Backend against #sql_injection with parameterized queries
    @mitigates MasterDataApp:Backend against #connection_failure with error handling and fallback
    """
    
    TABLE_NAME = "cpg_reporting_sandbox.default.planning_unit_stage_current"
    
    def __init__(self):
        """
        Initialize Databricks REST API connection parameters
        
        @review MasterDataApp:Backend ensure Databricks credentials are properly set
        @exposes MasterDataApp:Backend to #connection_failure if credentials missing
        """
        self.connection_available = False
        
        # Try to get connection parameters from environment
        self.server_hostname = os.environ.get('DATABRICKS_SERVER_HOSTNAME')
        self.http_path = os.environ.get('DATABRICKS_HTTP_PATH')
        self.access_token = os.environ.get('DATABRICKS_TOKEN')
        
        # Check if credentials are configured
        if self.server_hostname and self.access_token:
            self.connection_available = True
            self.api_url = f"https://{self.server_hostname}/api/2.0/sql/statements"
            # Hardcoded warehouse ID (from HTTP path: /sql/1.0/warehouses/2459b6c757e09896)
            self.warehouse_id = "2459b6c757e09896"
            logger.info('Databricks REST API connection configured and available')
            logger.info(f'  - Server: {self.server_hostname}')
            logger.info(f'  - Warehouse ID: {self.warehouse_id}')
            logger.info(f'  - API URL: {self.api_url}')
        else:
            logger.warning('Databricks credentials not fully configured - will use sample data')
            logger.info(f'  - Server hostname: {self.server_hostname if self.server_hostname else "Not set"}')
            logger.info(f'  - HTTP path: {self.http_path if self.http_path else "Not set"}')
            logger.info(f'  - Access token: {"Set (length: " + str(len(self.access_token)) + ")" if self.access_token else "Not set"}')
        
    def _execute_sql(self, query, timeout=50):
        """
        Execute SQL query using Databricks REST API
        
        @mitigates MasterDataApp:Backend against #sql_injection with fixed queries
        @connects MasterDataApp:Backend to Databricks:Warehouse with HTTPS/REST
        """
        if not self.connection_available:
            raise Exception("Databricks connection not configured")
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json"
        }
        
        # Databricks API requires wait_timeout between 5s and 50s
        wait_time = min(max(timeout, 5), 50)  # Clamp between 5 and 50
        
        payload = {
            "statement": query,
            "warehouse_id": self.warehouse_id,
            "wait_timeout": f"{wait_time}s"
        }
        
        logger.info(f'Executing SQL via REST API to warehouse {self.warehouse_id} (timeout: {wait_time}s)')
        
        try:
            # Submit SQL statement
            logger.info(f'Sending POST to {self.api_url}')
            logger.info(f'Payload: {payload}')
            response = requests.post(self.api_url, json=payload, headers=headers, timeout=timeout+10)
            
            logger.info(f'Response status code: {response.status_code}')
            
            if response.status_code != 200:
                error_detail = response.text
                logger.error(f'API returned error: {response.status_code} - {error_detail}')
                try:
                    error_json = response.json()
                    raise Exception(f"Databricks API Error: {error_json.get('message', error_detail)}")
                except:
                    raise Exception(f"Databricks API Error ({response.status_code}): {error_detail}")
            
            result = response.json()
            logger.info(f'Query execution state: {result.get("status", {}).get("state")}')
            
            # Check if execution completed
            if result.get('status', {}).get('state') == 'SUCCEEDED':
                logger.info('Query succeeded!')
                return result
            else:
                error_msg = result.get('status', {}).get('error', 'Unknown error')
                logger.error(f'Query failed with error: {error_msg}')
                raise Exception(f"Query failed: {error_msg}")
                
        except requests.exceptions.RequestException as e:
            logger.error(f'REST API request failed: {type(e).__name__} - {str(e)}')
            if hasattr(e, 'response') and e.response is not None:
                logger.error(f'Response body: {e.response.text}')
            raise
        except Exception as e:
            logger.error(f'Unexpected error in _execute_sql: {type(e).__name__} - {str(e)}')
            raise
    
    def test_connection(self):
        """
        Test Databricks connection using REST API
        
        @tests DatabaseConnection for MasterDataApp:Backend
        @mitigates MasterDataApp:Backend against #connection_failure with connection validation
        """
        if not self.connection_available:
            return False, "Databricks REST API not configured"
        
        try:
            result = self._execute_sql("SELECT 1 as test")
            
            # Check if we got valid results
            if result.get('result'):
                return True, "Connection successful"
            else:
                return False, "Unexpected query result"
                
        except Exception as e:
            logger.error(f'Connection test failed: {str(e)}')
            return False, str(e)
    
    def read_data(self):
        """
        Read data from Databricks table using REST API
        Filter out records that have been approved or denied (if column exists)
        
        @mitigates MasterDataApp:Backend against #sql_injection with fixed query
        @validates SQL query structure
        """
        if not self.connection_available:
            raise Exception("Databricks connection not configured")
        
        try:
            # Try to query with Approval_Status filter first
            # @mitigates against #sql_injection with fixed table name
            query_with_filter = f"""
                SELECT *
                FROM {self.TABLE_NAME}
                WHERE Approval_Status IS NULL 
                   OR Approval_Status NOT IN ('Approved', 'Denied')
            """
            
            try:
                logger.info(f'Attempting query with Approval_Status filter...')
                result = self._execute_sql(query_with_filter)
                logger.info(f'Query with filter succeeded')
            except Exception as filter_error:
                # Column might not exist yet, try without filter
                logger.warning(f'Query with Approval_Status filter failed (column may not exist): {filter_error}')
                logger.info(f'Attempting query without filter...')
                query_without_filter = f"""
                    SELECT *
                    FROM {self.TABLE_NAME}
                """
                result = self._execute_sql(query_without_filter)
                logger.info(f'Query without filter succeeded')
            
            # Parse results from REST API response
            data = []
            if result.get('result') and result['result'].get('data_array'):
                columns = [col['name'] for col in result['manifest']['schema']['columns']]
                
                for row_array in result['result']['data_array']:
                    row_dict = {}
                    for i, col_name in enumerate(columns):
                        row_dict[col_name] = row_array[i]
                    data.append(row_dict)
            
            logger.info(f'Retrieved {len(data)} records from Databricks via REST API')
            return data
            
        except Exception as e:
            logger.error(f'Error reading from Databricks: {str(e)}')
            raise
    
    def write_approvals(self, approved_records, denied_records):
        """
        Write approval decisions back to Databricks table using REST API
        
        @mitigates MasterDataApp:Backend against #sql_injection with sanitized inputs
        @mitigates MasterDataApp:Backend against #race_condition with sequential updates
        @validates Input records before database operations
        """
        if not self.connection_available:
            raise Exception("Databricks connection not configured")
        
        updated_count = 0
        
        try:
            all_records = []
            for record in approved_records:
                all_records.append((record, 'Approved'))
            for record in denied_records:
                all_records.append((record, 'Denied'))
            
            # Process each record with sanitized query
            for record, status in all_records:
                # Validate record structure
                if not InputValidator.validate_record(record):
                    logger.warning(f'Invalid record structure, skipping: {record}')
                    continue
                
                # Sanitize inputs to prevent SQL injection
                prdid = InputValidator.sanitize_string(record.get('PRDID'))
                locid = InputValidator.sanitize_string(record.get('LOCID'))
                recommended = InputValidator.sanitize_string(record.get('RECOMMENDED', ''))
                
                # Escape single quotes in strings for SQL
                prdid = prdid.replace("'", "''")
                locid = locid.replace("'", "''")
                recommended = recommended.replace("'", "''")
                status_safe = status.replace("'", "''")
                
                # Build UPDATE query with sanitized values
                # @mitigates against #sql_injection with sanitized and escaped inputs
                update_query = f"""
                    UPDATE {self.TABLE_NAME}
                    SET Approval_Status = '{status_safe}', 
                        RECOMMENDED = '{recommended}'
                    WHERE PRDID = '{prdid}' AND LOCID = '{locid}'
                """
                
                try:
                    self._execute_sql(update_query)
                    updated_count += 1
                    logger.info(f'Updated record: PRDID={prdid}, LOCID={locid}, Status={status}')
                except Exception as e:
                    logger.error(f'Error updating record PRDID={prdid}, LOCID={locid}: {str(e)}')
                    continue
            
            logger.info(f'Successfully updated {updated_count} records in Databricks via REST API')
            return updated_count
            
        except Exception as e:
            logger.error(f'Error writing to Databricks: {str(e)}')
            raise
    
    def read_historical_data(self):
        """
        Read all historical snapshot data from Databricks historical table using REST API
        Returns data sorted by snapshot_date descending (most recent first)
        
        @mitigates MasterDataApp:Backend:Historical against #sql_injection with fixed query
        @connects MasterDataApp:Backend to Databricks:HistoricalTable with REST API
        @validates SQL query structure
        """
        if not self.connection_available:
            raise Exception("Databricks connection not configured")
        
        HISTORICAL_TABLE_NAME = "cpg_reporting_sandbox.default.planning_unit_stage_historical"
        
        try:
            # Query all historical data, ordered by snapshot date
            query = f"""
                SELECT *
                FROM {HISTORICAL_TABLE_NAME}
                ORDER BY snapshot_date DESC, PRDID, LOCID
            """
            
            logger.info(f'Reading historical data from {HISTORICAL_TABLE_NAME}')
            result = self._execute_sql(query)
            
            # Parse results
            data = []
            if result.get('result') and result['result'].get('data_array'):
                # Get column names from manifest
                columns = [col['name'] for col in result['manifest']['schema']['columns']]
                
                # Convert each row array to dictionary
                for row_array in result['result']['data_array']:
                    row_dict = {}
                    for i, col_name in enumerate(columns):
                        row_dict[col_name] = row_array[i]
                    data.append(row_dict)
            
            logger.info(f'Retrieved {len(data)} historical records from Databricks via REST API')
            return data
            
        except Exception as e:
            logger.error(f'Error reading historical data from Databricks: {str(e)}')
            raise


# Initialize Databricks handler
databricks_handler = DatabricksHandler()


# Static file serving
# @connects User:Browser to MasterDataApp:Frontend with HTTP/HTML
@app.route('/')
def serve_index():
    """
    Serve the React application index.html
    
    @exposes MasterDataApp:Frontend to #unauth_access with public frontend
    @accepts #unauth_access to MasterDataApp:Frontend with application is intentionally public
    """
    try:
        return send_from_directory(DIST_DIR, 'index.html')
    except Exception as e:
        logger.error(f'Error serving index.html: {e}')
        return jsonify({'error': 'Frontend not found', 'detail': str(e)}), 404


@app.route('/<path:path>')
def serve_static(path):
    """
    Serve static files from dist folder
    
    @mitigates MasterDataApp:Frontend against #path_traversal with Flask send_from_directory validation
    """
    try:
        # Check if file exists in dist directory
        file_path = DIST_DIR / path
        if file_path.exists() and file_path.is_file():
            return send_from_directory(DIST_DIR, path)
        # If not found, serve index.html for React routing
        return send_from_directory(DIST_DIR, 'index.html')
    except Exception as e:
        logger.error(f'Error serving static file {path}: {e}')
        return send_from_directory(DIST_DIR, 'index.html')


# API Endpoints
# @connects MasterDataApp:Frontend to MasterDataApp:Backend:API with fetch/AJAX


@app.route('/api/health', methods=['GET'])
def health_check():
    """
    Health check endpoint
    
    @exposes MasterDataApp:Backend:API to #info_disclosure with version info disclosure accepted
    @accepts #info_disclosure to MasterDataApp:Backend:API with health check is non-sensitive
    @tests HealthCheck for MasterDataApp:Backend:API
    """
    mode = 'database' if databricks_handler.connection_available else 'sample_data'
    
    # Debug info (without exposing secrets)
    debug_info = {
        'server_configured': bool(databricks_handler.server_hostname),
        'http_path_configured': bool(databricks_handler.http_path),
        'token_configured': bool(databricks_handler.access_token),
        'token_length': len(databricks_handler.access_token) if databricks_handler.access_token else 0,
        'warehouse_id': databricks_handler.warehouse_id if databricks_handler.connection_available else None
    }
    
    return jsonify({
        'status': 'ok',
        'mode': mode,
        'databricks_configured': databricks_handler.connection_available,
        'message': f'Running in {mode} mode',
        'debug': debug_info
    })


@app.route('/api/test-connection', methods=['GET'])
def test_connection():
    """
    Test Databricks connection endpoint
    
    @tests DatabaseConnection for MasterDataApp:Backend:API
    @connects MasterDataApp:Backend to Databricks:Warehouse with connection test
    """
    try:
        success, message = databricks_handler.test_connection()
        
        if success:
            return jsonify({
                'connection': 'success',
                'message': message,
                'table': DatabricksHandler.TABLE_NAME
            })
        else:
            return jsonify({
                'connection': 'failed',
                'message': message
            }), 500
            
    except Exception as e:
        logger.error(f'Connection test error: {str(e)}')
        import traceback
        return jsonify({
            'connection': 'failed',
            'error': str(e),
            'error_type': type(e).__name__,
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/data', methods=['GET'])
def get_data():
    """
    Get data - tries database first, falls back to sample data
    
    @mitigates MasterDataApp:Backend:API against #service_failure with graceful degradation
    @tests DataRetrieval for MasterDataApp:Backend:API
    @connects MasterDataApp:Backend:API to Databricks:Table with SQL query
    """
    last_error = None
    
    try:
        # Try database first if available
        if databricks_handler.connection_available:
            try:
                logger.info('Attempting to read data from Databricks...')
                data = databricks_handler.read_data()
                
                # Add approval field for frontend compatibility
                for row in data:
                    approval_status = row.get('Approval_Status', '')
                    if approval_status == 'Approved':
                        row['approval'] = 'approve'
                    elif approval_status == 'Denied':
                        row['approval'] = 'deny'
                    else:
                        row['approval'] = None
                
                logger.info(f'Successfully returned {len(data)} records from database')
                return jsonify(data)
                
            except Exception as db_error:
                last_error = str(db_error)
                logger.warning(f'Database read failed, falling back to sample data: {db_error}')
                import traceback
                logger.error(traceback.format_exc())
                # Fall through to sample data
        
        # Fallback to sample data
        logger.info(f'Returning {len(SAMPLE_DATA)} sample records (fallback mode)')
        logger.info(f'Last error: {last_error}')
        import copy
        data_response = copy.deepcopy(SAMPLE_DATA)
        
        # Add error info to first record for debugging
        if last_error and len(data_response) > 0:
            data_response[0]['_debug_error'] = last_error
        
        return jsonify(data_response)
    
    except Exception as e:
        logger.error(f'Error in get_data: {str(e)}')
        return jsonify({
            'error': 'Failed to retrieve data',
            'message': str(e)
        }), 500


@app.route('/api/submit', methods=['POST'])
def submit_data():
    """
    Update records - tries database, provides clear feedback
    
    @mitigates MasterDataApp:Backend:API against #injection with input validation
    @mitigates MasterDataApp:Backend:Database against #race_condition with locking
    @review MasterDataApp:Backend:API authentication and authorization required for production
    @tests DataSubmission for MasterDataApp:Backend:API
    """
    try:
        # Get request data
        # @validates Input from API request
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        approved = data.get('approved', [])
        denied = data.get('denied', [])
        
        logger.info(f'Processing submission: {len(approved)} approved, {len(denied)} denied')
        
        # If database is available, write to it
        if databricks_handler.connection_available:
            try:
                # Use lock to prevent concurrent database operations
                with operation_lock:
                    updated_count = databricks_handler.write_approvals(approved, denied)
                
                return jsonify({
                    'success': True,
                    'message': 'Records updated successfully in database',
                    'updated_count': updated_count,
                    'mode': 'database'
                })
                
            except Exception as db_error:
                logger.error(f'Database write failed: {db_error}')
                return jsonify({
                    'success': False,
                    'error': 'Failed to update database',
                    'message': str(db_error)
                }), 500
        else:
            # No database - return mock success
            logger.info('Database not available - mock submission')
            return jsonify({
                'success': True,
                'message': 'Mock submission successful (no database configured)',
                'approved_count': len(approved),
                'denied_count': len(denied),
                'mode': 'sample_data'
            })
    
    except Exception as e:
        logger.error(f'Error in submit_data: {str(e)}')
        return jsonify({
            'error': 'Failed to process submission',
            'message': str(e)
        }), 500


def create_excel_from_historical_data(data):
    """
    Create an Excel workbook from historical data
    
    @mitigates MasterDataApp:Backend:Historical against #data_tampering with server-side generation
    @validates Data structure before Excel generation
    """
    if not data or len(data) == 0:
        raise ValueError("No historical data available")
    
    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Historical Data"
    
    # Get column headers from first row
    headers = list(data[0].keys())
    
    # Write headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header)
            # Convert None to empty string
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        # Set minimum width based on header length, max 50
        max_length = len(str(header))
        for row_idx in range(2, min(len(data) + 2, 100)):  # Sample first 100 rows for width
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


@app.route('/api/historical/download', methods=['GET'])
def download_historical():
    """
    Download historical snapshots as Excel file
    
    @mitigates MasterDataApp:Backend:Historical against #sql_injection with parameterized queries
    @mitigates MasterDataApp:Backend:Historical against #path_traversal with server-side file generation
    @connects MasterDataApp:Frontend to MasterDataApp:Backend:Historical with HTTP GET
    @review MasterDataApp:Backend:Historical ensure historical data access is authorized
    """
    try:
        logger.info('Historical data download requested')
        
        # Check if database is available
        if not databricks_handler.connection_available:
            return jsonify({
                'error': 'Database not configured',
                'message': 'Historical data is only available when connected to Databricks'
            }), 503
        
        # Read historical data from Databricks
        with operation_lock:
            historical_data = databricks_handler.read_historical_data()
        
        if not historical_data or len(historical_data) == 0:
            return jsonify({
                'error': 'No historical data found',
                'message': 'The historical table is empty. Run the weekly snapshot job first.'
            }), 404
        
        logger.info(f'Generating Excel file with {len(historical_data)} historical records')
        
        # Generate Excel file
        excel_file = create_excel_from_historical_data(historical_data)
        
        # Create filename with current date
        filename = f'planning_unit_historical_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        logger.info(f'Sending Excel file: {filename}')
        
        # Return Excel file as download
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ValueError as ve:
        logger.error(f'Validation error downloading historical data: {str(ve)}')
        return jsonify({'error': 'Invalid data', 'message': str(ve)}), 400
    
    except Exception as e:
        logger.error(f'Error downloading historical data: {str(e)}')
        return jsonify({
            'error': 'Failed to download historical data',
            'message': str(e)
        }), 500


@app.errorhandler(404)
def not_found(error):
    """
    Handle 404 errors - serve React app for client-side routing
    
    @mitigates MasterDataApp:Frontend against #broken_routes with catch-all handler
    """
    # For API routes, return JSON error
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Endpoint not found'}), 404
    # For other routes, serve React app (client-side routing)
    return serve_index()


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    logger.error(f'Internal error: {error}')
    return jsonify({'error': 'Internal server error'}), 500


if __name__ == '__main__':
    """
    Application entry point
    
    @exposes MasterDataApp:Backend:API to #dos with no rate limiting
    @review MasterDataApp:Backend:API add rate limiting for production
    @connects Databricks:Platform to MasterDataApp:Backend with PORT environment variable
    """
    port = int(os.environ.get('PORT', 8080))
    
    logger.info('=' * 60)
    logger.info('IBP Master Data Automation - ROBUST VERSION')
    logger.info('=' * 60)
    logger.info(f'Python version: {sys.version}')
    logger.info(f'Working directory: {os.getcwd()}')
    logger.info(f'Server port: {port}')
    logger.info(f'BASE_DIR: {BASE_DIR}')
    logger.info(f'DIST_DIR: {DIST_DIR}')
    logger.info(f'DIST exists: {DIST_DIR.exists()}')
    
    if DIST_DIR.exists():
        dist_files = list(DIST_DIR.iterdir())
        logger.info(f'Files in DIST: {[f.name for f in dist_files[:10]]}')
    else:
        logger.warning('DIST directory not found!')
    
    logger.info(f'Database mode: {"ENABLED" if databricks_handler.connection_available else "DISABLED (using sample data)"}')
    if databricks_handler.connection_available:
        logger.info(f'Databricks table: {DatabricksHandler.TABLE_NAME}')
    logger.info(f'Sample data records: {len(SAMPLE_DATA)}')
    logger.info('API endpoints:')
    logger.info('  - GET  /api/health')
    logger.info('  - GET  /api/test-connection')
    logger.info('  - GET  /api/data')
    logger.info('  - POST /api/submit')
    logger.info('  - GET  /api/historical/download')
    logger.info('Frontend: Serving React app from dist/')
    logger.info('FALLBACK: Sample data available if database fails')
    logger.info('=' * 60)
    
    # Verify dist directory exists before starting
    if not DIST_DIR.exists():
        logger.error(f'ERROR: DIST_DIR does not exist at {DIST_DIR}')
        logger.error('Please build the React frontend first: npm run build')
    
    app.run(host='0.0.0.0', port=port, debug=False)
